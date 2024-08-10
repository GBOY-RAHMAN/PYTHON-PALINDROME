# question 1

import openpyxl
from openpyxl import Workbook

# Create a workbook
wb = Workbook()

# Create a worksheet
ws = wb.active  # By default, a new workbook has one active sheet
ws.title = "Sheet1"

# Fill the cells from A1 to C2 in the file with data
data = [
    [9, 7, 3],
    [1, 5, 4]
]

for i, row in enumerate(data, start=1):
    for j, value in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=value)

# Create a new sheet
ws2 = wb.create_sheet(title="pythonExcel")

# Fill the new sheet with any data
custom_data = [
    ["Student_Name", "Age", "City"],
    ["Gbolahan", 25, "Toronto"],
    ["jame", 18, "Los Angeles"],
    ["john", 35, "Chicago"]
]

for i, row in enumerate(custom_data, start=1):
    for j, value in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=value)

# Save the workbook to a file
wb.save("assigment.xlsx")



# question 2

# Write a Python function sum_columns() that reads an Excel file, computes the sum of
# each column in the first worksheet, and writes the sums to a new worksheet in the
# same file.
from openpyxl import load_workbook

def sum_columns(file_path):
    # Load the existing workbook
    wb = load_workbook("assigment.xlsx")
    
    # Select the first worksheet
    ws1 = wb.active
    
    # Calculate the sum of each column
    column_sums = {}
    for column in ws1.iter_cols(values_only=True):
        col_letter = column[0].column_letter
        column_sums[col_letter] = sum(cell for cell in column if isinstance(cell, (int, float)))
    
    # Create a new worksheet to store the sums
    ws_sum = wb.create_sheet(title="ColumnSums")
    
    # Write the sums to the new worksheet
    for i, (col, sum_value) in enumerate(column_sums.items(), start=1):
        ws_sum.cell(row=1, column=i, value=col)
        ws_sum.cell(row=2, column=i, value=sum_value)
    
    # Save the workbook
    wb.save("assigment.xlsx")

# Use the function
sum_columns("assigment.xlsx")